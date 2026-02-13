from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Dict, List, Tuple

from programmatic_content import OPTIONAL_PIPELINE_COLUMNS, REQUIRED_COLUMNS, build_entries


CORE_WRITE_COLUMNS = [
    'entry_type',
    'route_path',
    'slug',
    'title',
    'meta_description',
    'h1',
    'intro',
    'body',
    'primary_cta_label',
    'primary_cta_url',
    'secondary_cta_label',
    'secondary_cta_url',
    'generation_seed_prompt',
    'image_url',
    'feature_bullets',
    'faq_pairs',
    'status',
    'tags',
    'updated_at',
]


LIST_PIPE_FIELDS = {'feature_bullets', 'secondary_keywords'}
LIST_CSV_FIELDS = {'tags'}


def _safe_text(value: object) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _required_columns_present(headers: List[str]) -> List[str]:
    missing = REQUIRED_COLUMNS - set(headers)
    if not missing:
        return []
    return [f"Missing required columns: {', '.join(sorted(missing))}"]


def load_rows_with_headers(source_path: str, sheet_name: str = 'content') -> Tuple[List[Dict[str, str]], List[str], List[str]]:
    if not os.path.exists(source_path):
        return [], [], [f'Source spreadsheet not found: {source_path}']

    ext = Path(source_path).suffix.lower()
    if ext in {'.csv', '.tsv'}:
        delimiter = ',' if ext == '.csv' else '\t'
        with open(source_path, 'r', encoding='utf-8-sig', newline='') as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            if not reader.fieldnames:
                return [], [], ['Spreadsheet has no header row']

            headers = [_safe_text(h) for h in reader.fieldnames if _safe_text(h)]
            missing_errors = _required_columns_present(headers)
            if missing_errors:
                return [], headers, missing_errors

            rows: List[Dict[str, str]] = []
            for row_number, row in enumerate(reader, start=2):
                row_data: Dict[str, str] = {'_row_number': str(row_number)}
                for header in headers:
                    row_data[header] = _safe_text((row or {}).get(header, ''))
                rows.append(row_data)
            return rows, headers, []

    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            return [], [], ['XLSX input requires openpyxl. Install with pip install openpyxl']

        workbook = load_workbook(filename=source_path)
        if sheet_name not in workbook.sheetnames:
            return [], [], [f"Sheet '{sheet_name}' not found. Available: {', '.join(workbook.sheetnames)}"]

        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        if max_row < 1:
            return [], [], ['Spreadsheet is empty']

        headers = []
        for col in range(1, max_col + 1):
            header = _safe_text(sheet.cell(row=1, column=col).value)
            if header:
                headers.append(header)
        if not headers:
            return [], [], ['Spreadsheet has no header row']

        missing_errors = _required_columns_present(headers)
        if missing_errors:
            return [], headers, missing_errors

        rows: List[Dict[str, str]] = []
        for row_number in range(2, max_row + 1):
            row_data: Dict[str, str] = {'_row_number': str(row_number)}
            has_data = False
            for col, header in enumerate(headers, start=1):
                value = _safe_text(sheet.cell(row=row_number, column=col).value)
                if value:
                    has_data = True
                row_data[header] = value
            if has_data:
                rows.append(row_data)

        return rows, headers, []

    return [], [], [f'Unsupported spreadsheet extension: {ext}. Use .csv, .tsv, or .xlsx']


def ensure_headers(headers: List[str]) -> List[str]:
    ordered = []
    seen = set()

    for key in CORE_WRITE_COLUMNS:
        if key not in seen:
            ordered.append(key)
            seen.add(key)

    for key in OPTIONAL_PIPELINE_COLUMNS:
        if key not in seen:
            ordered.append(key)
            seen.add(key)

    for key in headers:
        k = _safe_text(key)
        if k and k not in seen and k != '_row_number':
            ordered.append(k)
            seen.add(k)

    return ordered


def _join_faq(items: List[Dict[str, str]]) -> str:
    chunks = []
    for item in items:
        question = _safe_text(item.get('question'))
        answer = _safe_text(item.get('answer'))
        if not question:
            continue
        chunks.append(f'{question}::{answer}')
    return '||'.join(chunks)


def _list_to_pipe(values: List[str]) -> str:
    return '|'.join(_safe_text(item) for item in values if _safe_text(item))


def _list_to_csv(values: List[str]) -> str:
    return ','.join(_safe_text(item) for item in values if _safe_text(item))


def normalize_entry_from_row(row: Dict[str, str]) -> Tuple[Dict[str, object] | None, List[str]]:
    payload = dict(row)
    payload.setdefault('_row_number', row.get('_row_number', '?'))
    entries, errors = build_entries([payload])
    if errors or not entries:
        return None, errors or ['Unknown entry normalization error']
    return entries[0], []


def apply_entry_to_row(row: Dict[str, str], entry: Dict[str, object]) -> Dict[str, str]:
    updated = dict(row)

    for key in [
        'entry_type',
        'route_path',
        'slug',
        'title',
        'meta_description',
        'h1',
        'intro',
        'body',
        'primary_cta_label',
        'primary_cta_url',
        'secondary_cta_label',
        'secondary_cta_url',
        'generation_seed_prompt',
        'image_url',
        'status',
        'updated_at',
        'content_status',
        'image_status',
        'primary_keyword',
        'content_brief',
        'image_style',
        'image_aspect_ratio',
        'image_prompt_override',
        'asset_local_path',
        'asset_hash',
        'generation_batch_id',
        'last_generated_at',
        'last_reviewed_at',
        'qa_notes',
    ]:
        updated[key] = _safe_text(entry.get(key, updated.get(key, '')))

    updated['feature_bullets'] = _list_to_pipe(list(entry.get('feature_bullets') or []))
    updated['secondary_keywords'] = _list_to_pipe(list(entry.get('secondary_keywords') or []))
    updated['tags'] = _list_to_csv(list(entry.get('tags') or []))
    updated['faq_pairs'] = _join_faq(list(entry.get('faq') or []))
    return updated


def save_rows(source_path: str, sheet_name: str, rows: List[Dict[str, str]], headers: List[str]) -> None:
    ext = Path(source_path).suffix.lower()
    headers = [h for h in headers if h and h != '_row_number']

    if ext in {'.csv', '.tsv'}:
        delimiter = ',' if ext == '.csv' else '\t'
        with open(source_path, 'w', encoding='utf-8', newline='') as handle:
            writer = csv.DictWriter(handle, fieldnames=headers, delimiter=delimiter)
            writer.writeheader()
            for row in rows:
                writer.writerow({header: _safe_text(row.get(header, '')) for header in headers})
        return

    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            raise RuntimeError('XLSX output requires openpyxl. Install with pip install openpyxl')

        workbook = load_workbook(source_path)
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook")

        sheet = workbook[sheet_name]

        existing_headers = []
        for col in range(1, sheet.max_column + 1):
            val = _safe_text(sheet.cell(row=1, column=col).value)
            if val:
                existing_headers.append(val)

        combined_headers = []
        seen = set()
        for item in headers + existing_headers:
            val = _safe_text(item)
            if val and val not in seen:
                seen.add(val)
                combined_headers.append(val)

        for col, header in enumerate(combined_headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        max_existing_rows = max(sheet.max_row, len(rows) + 1)
        for row_idx in range(2, max_existing_rows + 1):
            payload = rows[row_idx - 2] if row_idx - 2 < len(rows) else {}
            for col_idx, header in enumerate(combined_headers, start=1):
                value = _safe_text(payload.get(header, ''))
                sheet.cell(row=row_idx, column=col_idx, value=value if value else None)

        workbook.save(source_path)
        return

    raise RuntimeError(f'Unsupported spreadsheet extension: {ext}')

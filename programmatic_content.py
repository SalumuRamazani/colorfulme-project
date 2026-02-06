"""Programmatic page/tool content pipeline.

This module provides:
- Spreadsheet ingestion from CSV or XLSX
- Validation and normalization of content rows
- Manifest generation for runtime routing
"""
from __future__ import annotations

import csv
import json
import os
from datetime import datetime, timezone
from typing import Dict, List, Tuple

DEFAULT_SOURCE_PATH = "content/programmatic_content.csv"
DEFAULT_SHEET_NAME = "content"
DEFAULT_MANIFEST_PATH = "static/data/programmatic_content_manifest.json"

ALLOWED_ENTRY_TYPES = {"page", "tool"}
PUBLISHED_STATUSES = {"published", "live", "active"}

REQUIRED_COLUMNS = {
    "entry_type",
    "route_path",
    "title",
}


def _safe_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_path(raw_path: str) -> str:
    value = _safe_text(raw_path)
    if not value:
        return ""
    if not value.startswith("/"):
        value = "/" + value
    if len(value) > 1 and value.endswith("/"):
        value = value[:-1]
    return value


def _normalize_entry_type(raw_type: str) -> str:
    entry_type = _safe_text(raw_type).lower()
    return entry_type


def _normalize_status(raw_status: str) -> str:
    status = _safe_text(raw_status).lower()
    return status or "published"


def _split_pipe_list(value: str) -> List[str]:
    cleaned = _safe_text(value)
    if not cleaned:
        return []
    return [part.strip() for part in cleaned.split("|") if part.strip()]


def _split_csv_list(value: str) -> List[str]:
    cleaned = _safe_text(value)
    if not cleaned:
        return []
    return [part.strip() for part in cleaned.split(",") if part.strip()]


def _parse_faq_pairs(raw_faq_pairs: str) -> List[Dict[str, str]]:
    cleaned = _safe_text(raw_faq_pairs)
    if not cleaned:
        return []

    faq_items: List[Dict[str, str]] = []
    for chunk in cleaned.split("||"):
        part = chunk.strip()
        if not part:
            continue

        if "::" in part:
            question, answer = part.split("::", 1)
        else:
            question, answer = part, ""

        question = question.strip()
        answer = answer.strip()
        if question:
            faq_items.append({"question": question, "answer": answer})

    return faq_items


def _split_paragraphs(text: str) -> List[str]:
    cleaned = _safe_text(text)
    if not cleaned:
        return []
    return [paragraph.strip() for paragraph in cleaned.split("\n\n") if paragraph.strip()]


def _read_csv_rows(source_path: str) -> Tuple[List[Dict[str, str]], List[str]]:
    rows: List[Dict[str, str]] = []
    errors: List[str] = []

    with open(source_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return [], ["Spreadsheet has no header row"]

        missing_columns = REQUIRED_COLUMNS - set(name.strip() for name in reader.fieldnames)
        if missing_columns:
            return [], [f"Missing required columns: {', '.join(sorted(missing_columns))}"]

        for row_number, row in enumerate(reader, start=2):
            rows.append({"_row_number": str(row_number), **{k: _safe_text(v) for k, v in row.items()}})

    return rows, errors


def _read_xlsx_rows(source_path: str, sheet_name: str) -> Tuple[List[Dict[str, str]], List[str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return [], [
            "XLSX input requires openpyxl. Install it with: pip install openpyxl"
        ]

    workbook = load_workbook(filename=source_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return [], [
            f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
        ]

    sheet = workbook[sheet_name]
    row_iterator = sheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iterator)
    except StopIteration:
        return [], ["Spreadsheet is empty"]

    headers = [_safe_text(value) for value in header_row]
    missing_columns = REQUIRED_COLUMNS - set(headers)
    if missing_columns:
        return [], [f"Missing required columns: {', '.join(sorted(missing_columns))}"]

    rows: List[Dict[str, str]] = []
    for row_number, values in enumerate(row_iterator, start=2):
        row_dict: Dict[str, str] = {"_row_number": str(row_number)}
        for index, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = _safe_text(values[index] if index < len(values) else "")
        rows.append(row_dict)

    return rows, []


def read_spreadsheet_rows(source_path: str, sheet_name: str = DEFAULT_SHEET_NAME) -> Tuple[List[Dict[str, str]], List[str]]:
    if not os.path.exists(source_path):
        return [], [f"Source spreadsheet not found: {source_path}"]

    extension = os.path.splitext(source_path)[1].lower()
    if extension in {".csv", ".tsv"}:
        if extension == ".tsv":
            # Convert TSV into CSV reader-friendly format by reusing csv with delimiter.
            rows: List[Dict[str, str]] = []
            with open(source_path, "r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle, delimiter="\t")
                if not reader.fieldnames:
                    return [], ["Spreadsheet has no header row"]

                missing_columns = REQUIRED_COLUMNS - set(name.strip() for name in reader.fieldnames)
                if missing_columns:
                    return [], [f"Missing required columns: {', '.join(sorted(missing_columns))}"]

                for row_number, row in enumerate(reader, start=2):
                    rows.append({"_row_number": str(row_number), **{k: _safe_text(v) for k, v in row.items()}})

            return rows, []

        return _read_csv_rows(source_path)

    if extension == ".xlsx":
        return _read_xlsx_rows(source_path, sheet_name)

    return [], [f"Unsupported spreadsheet extension: {extension}. Use .csv, .tsv, or .xlsx"]


def build_entries(rows: List[Dict[str, str]]) -> Tuple[List[Dict[str, object]], List[str]]:
    entries: List[Dict[str, object]] = []
    errors: List[str] = []
    seen_paths = set()

    for row in rows:
        row_number = row.get("_row_number", "?")
        entry_type = _normalize_entry_type(row.get("entry_type", ""))
        route_path = _normalize_path(row.get("route_path", ""))
        title = _safe_text(row.get("title", ""))

        if not entry_type:
            errors.append(f"Row {row_number}: entry_type is required")
            continue
        if entry_type not in ALLOWED_ENTRY_TYPES:
            errors.append(f"Row {row_number}: entry_type '{entry_type}' must be one of: {', '.join(sorted(ALLOWED_ENTRY_TYPES))}")
            continue
        if not route_path or route_path == "/":
            errors.append(f"Row {row_number}: route_path must be a non-root URL path")
            continue
        if not title:
            errors.append(f"Row {row_number}: title is required")
            continue
        if route_path in seen_paths:
            errors.append(f"Row {row_number}: duplicate route_path '{route_path}'")
            continue

        seen_paths.add(route_path)

        intro = _safe_text(row.get("intro", ""))
        body = _safe_text(row.get("body", ""))

        entry = {
            "entry_type": entry_type,
            "route_path": route_path,
            "slug": _safe_text(row.get("slug", "")) or route_path.strip("/").replace("/", "-"),
            "title": title,
            "meta_description": _safe_text(row.get("meta_description", "")),
            "h1": _safe_text(row.get("h1", "")) or title,
            "intro": intro,
            "intro_paragraphs": _split_paragraphs(intro),
            "body": body,
            "body_paragraphs": _split_paragraphs(body),
            "primary_cta_label": _safe_text(row.get("primary_cta_label", "")),
            "primary_cta_url": _safe_text(row.get("primary_cta_url", "")),
            "secondary_cta_label": _safe_text(row.get("secondary_cta_label", "")),
            "secondary_cta_url": _safe_text(row.get("secondary_cta_url", "")),
            "generator_slug": _safe_text(row.get("generator_slug", "")),
            "template_slug": _safe_text(row.get("template_slug", "")),
            "feature_bullets": _split_pipe_list(row.get("feature_bullets", "")),
            "faq": _parse_faq_pairs(row.get("faq_pairs", "")),
            "tags": _split_csv_list(row.get("tags", "")),
            "status": _normalize_status(row.get("status", "")),
            "updated_at": _safe_text(row.get("updated_at", "")),
        }

        entries.append(entry)

    return entries, errors


def build_manifest(entries: List[Dict[str, object]], source_path: str) -> Dict[str, object]:
    pages_count = sum(1 for entry in entries if entry.get("entry_type") == "page")
    tools_count = sum(1 for entry in entries if entry.get("entry_type") == "tool")

    return {
        "version": 1,
        "source_path": source_path,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "counts": {
            "total": len(entries),
            "pages": pages_count,
            "tools": tools_count,
        },
        "entries": entries,
    }


def generate_manifest_from_spreadsheet(
    source_path: str,
    output_path: str,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> Tuple[Dict[str, object], List[str]]:
    rows, read_errors = read_spreadsheet_rows(source_path, sheet_name=sheet_name)
    if read_errors:
        return {}, read_errors

    entries, validation_errors = build_entries(rows)
    if validation_errors:
        return {}, validation_errors

    manifest = build_manifest(entries, source_path=source_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as handle:
        json.dump(manifest, handle, indent=2)

    return manifest, []


def load_manifest(manifest_path: str) -> Dict[str, object]:
    if not os.path.exists(manifest_path):
        return {
            "version": 1,
            "source_path": "",
            "generated_at": "",
            "counts": {"total": 0, "pages": 0, "tools": 0},
            "entries": [],
        }

    with open(manifest_path, "r", encoding="utf-8") as handle:
        data = json.load(handle)

    if "entries" not in data or not isinstance(data["entries"], list):
        return {
            "version": 1,
            "source_path": manifest_path,
            "generated_at": "",
            "counts": {"total": 0, "pages": 0, "tools": 0},
            "entries": [],
        }

    return data


def build_published_route_index(manifest: Dict[str, object]) -> Dict[str, Dict[str, object]]:
    entries = manifest.get("entries", [])
    route_index: Dict[str, Dict[str, object]] = {}

    for entry in entries:
        if not isinstance(entry, dict):
            continue

        status = _normalize_status(entry.get("status", ""))
        if status not in PUBLISHED_STATUSES:
            continue

        route_path = _normalize_path(entry.get("route_path", ""))
        if not route_path or route_path == "/":
            continue

        route_index[route_path] = entry

    return route_index
